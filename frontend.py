import streamlit as st
import requests
from datetime import date, datetime
import io
import os
from openpyxl import load_workbook

# 設定要連線的後端 API 網址
API_URL = "https://hitachi-sr-system-backend.onrender.com"

st.set_page_config(page_title="日立維修服務系統", layout="wide") # 改為 wide 讓畫面更寬廣
st.title("🛠️ 日立維修服務報告書 (完整版)")

# --- 核心功能：產生 Excel 報表 ---
def generate_excel(form_data, result):
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(current_dir, "template.xlsx")
        
        wb = load_workbook(template_path)
        ws = wb.active 
        
        # ==========================================
        # ⚠️ 【座標修改區】 ⚠️ 
        # 請根據你 template.xlsx 中合併儲存格的「最左上角」座標修改以下引號內的代碼！
        # ==========================================
        
        # 1. 基本資訊
        ws["AQ3"] = str(form_data["簽發日"])        # 簽發日 (請確認座標)
        ws["AS6"] = form_data["SVO_NO"]            # SVO. NO.
        ws["H7"] = form_data["顧客名"]            # 顧客名
        ws["AS8"] = form_data["依賴作業NO"]        # 依賴作業NO.
        ws["H9"] = form_data["承辦人"]            # 承辦人
        
        # 2. 設備資訊
        ws["H12"] = form_data["品名"]             # 品名(Model)
        ws["X12"] = form_data["形式"]             # 形式
        ws["AN12"] = form_data["SN"]               # S/N
        ws["BC12"] = "〇" if form_data["保固狀態"] == "保證期間內" else "" # 保證期間内
        ws["BC13"] = "〇" if form_data["保固狀態"] == "保證期間外" else "" # 保證期間外
        ws["H14"] = str(form_data["製造年月"])    # 製造年月
        ws["X14"] = str(form_data["啟用年月"])    # 啟用年月
        ws["AN14"] = form_data["Tool_ID"]          # Tool ID
        ws["BC14"] = "〇" if form_data["完工狀態"] == "已完工" else "" # 已完工
        ws["BC15"] = "〇" if form_data["完工狀態"] == "未完工" else "" # 未完工
        
        # 3. 作業內容與細節
        ws["G16"] = form_data["依賴內容"]
        ws["G18"] = form_data["現象內容"]
        ws["Z18"] = form_data["原因內容"]
        ws["AS18"] = form_data["處置內容"]
        
        # 4. 其他資訊
        ws["B73"] = form_data["內部連絡事項"]
        ws["I76"] = form_data["作業區分"]
        ws["AB76"] = form_data["作業內容"]
        ws["AV76"] = form_data["Error_Message"]
        ws["AO79"] = form_data["承認人"]
        ws["AU79"] = form_data["審查人"]
        ws["BA79"] = form_data["製作人"]
        
        # 5. 時數計算結果 (從後端回傳的值)
        # 假設平日時數在 Q21，夜間在 Q23 (請修改成正確座標)
        ws["AR64"] = result["計算結果"]["平日時數"]
        ws["AR66"] = result["計算結果"]["夜間/假日時數"]
        
        # ==========================================
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except Exception as e:
        st.error(f"Excel 產生失敗: {e}")
        return None

# --- 前端表單介面 ---
with st.form("full_service_form"):
    
    st.subheader("1. 基本資料")
    col1, col2, col3 = st.columns(3)
    with col1:
        issue_date = st.date_input("簽發日", date.today())
        customer = st.selectbox("顧客名", ["Fab 8A", "Fab 8B", "Fab 12", "其他"])
    with col2:
        svo_no = st.text_input("SVO. NO.")
        request_no = st.text_input("依賴作業NO.")
    with col3:
        pic = st.text_input("承辦人")
        worker_name = st.text_input("工程師姓名 (系統紀錄用)")

    st.subheader("2. 設備資訊")
    col4, col5, col6, col7 = st.columns(4)
    with col4:
        model = st.text_input("品名 (Model)", "CD_SEM")
        mfg_date = st.date_input("製造年月", date(2020, 1, 1))
    with col5:
        model_type = st.text_input("形式")
        install_date = st.date_input("啟用年月", date(2020, 1, 1))
    with col6:
        sn = st.text_input("S/N (設備序號)")
        tool_id = st.text_input("Tool ID", "SR-001")
    with col7:
        warranty = st.radio("保固狀態", ["保證期間內", "保證期間外"])
        completion = st.radio("完工狀態", ["已完工", "未完工"])

    st.subheader("3. 處理內容紀錄")
    request_content = st.text_area("依賴內容", height=68)
    col8, col9, col10 = st.columns(3)
    with col8:
        phenomenon = st.text_area("現象內容", height=100)
    with col9:
        cause = st.text_area("原因內容", height=100)
    with col10:
        action = st.text_area("處置內容", height=100)
        
    st.subheader("4. 作業代碼與時間")
    col11, col12, col13 = st.columns(3)
    with col11:
        work_class = st.selectbox("作業區分", ["01:Contract for Maintenance", "02:Repair", "05:Integrated Preventive Maintenance", "其他"])
        start_time = st.time_input("作業開始時間")
    with col12:
        work_content = st.selectbox("作業內容", ["01:Contract for Maintenance", "02:Repair", "10:Application Support", "其他"])
        end_time = st.time_input("作業結束時間")
    with col13:
        error_msg = st.text_input("Error Message")
        is_workday = st.checkbox("是否為平日上班日？", value=True)

    st.subheader("5. 備註與簽核")
    internal_notes = st.text_area("內部連絡事項")
    col14, col15, col16 = st.columns(3)
    with col14:
        approver = st.text_input("承認人")
    with col15:
        reviewer = st.text_input("審查人")
    with col16:
        creator = st.text_input("製作人", worker_name) # 預設帶入工程師名字

    # 送出按鈕
    submitted = st.form_submit_button("1. 提交紀錄並產生報表", type="primary", use_container_width=True)

# --- 提交後的邏輯 ---
if submitted:
    start_dt = datetime.combine(issue_date, start_time).isoformat()
    end_dt = datetime.combine(issue_date, end_time).isoformat()
    
    # 傳給後端 API 的精簡資料 (為了計算時數與存檔)
    api_payload = {
        "date": str(issue_date),
        "worker_name": worker_name if worker_name else creator,
        "fab_name": customer,
        "tool_id": tool_id,
        "subject": request_content,
        "start_time": start_dt,
        "end_time": end_dt,
        "is_workday": is_workday
    }
    
    # 用來產生 Excel 的完整表單資料
    excel_payload = {
        "簽發日": issue_date,
        "SVO_NO": svo_no,
        "顧客名": customer,
        "依賴作業NO": request_no,
        "承辦人": pic,
        "品名": model,
        "形式": model_type,
        "SN": sn,
        "保固狀態": warranty,
        "製造年月": mfg_date,
        "啟用年月": install_date,
        "Tool_ID": tool_id,
        "完工狀態": completion,
        "依賴內容": request_content,
        "現象內容": phenomenon,
        "原因內容": cause,
        "處置內容": action,
        "內部連絡事項": internal_notes,
        "作業區分": work_class,
        "作業內容": work_content,
        "Error_Message": error_msg,
        "承認人": approver,
        "審查人": reviewer,
        "製作人": creator
    }
    
    try:
        # 打 API 取得計算結果
        response = requests.post(f"{API_URL}/api/records", json=api_payload)
        if response.status_code == 200:
            result = response.json()
            st.success("✅ 系統紀錄新增成功！時數已計算完成。")
            
            # 將資料與計算結果結合，產生 Excel
            excel_data = generate_excel(excel_payload, result)
            
            if excel_data:
                st.download_button(
                    label="2. 📥 下載正式 Excel 服務報告書",
                    data=excel_data,
                    file_name=f"Service_Report_{tool_id}_{issue_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.error("❌ 送出失敗，請確認後端連線。")
    except Exception as e:
        st.error(f"連線錯誤: {e}")